#!/usr/bin/env python
# -*- coding: utf-8 -*-

""" batch_rion52.py
+++++++++++++++++++
©joaquinstevens
Santiago, Chile.
April 2016.

Notes:
- This script calculates global descriptors Leq, Lmax, Lmin per file of an NL-52 sound level meter.
- The measurement file .rnd exported from the equipment is the input to the calculation routine.
- This routine calculates the descriptors for all .nrd files in the working folder.
- These data are finally written in an Excel 2010 spreadsheet for use in evaluation reports.
- Rion NL52 datasheet: http://bit.ly/2lHtvrh

"""

import numpy as np              
import openpyxl as pyxl         
import glob
import string

# Generates list names files.rnd in folder.
nombre = []
for filename in glob.glob('*.csv'):
    nombre.append(filename)

# Generate data array to store total results in folder.
LEQ = np.zeros((len(nombre),34))
LMAX = np.zeros((len(nombre),34))
LMIN = np.zeros((len(nombre),34))
D = np.zeros(len(nombre))

# Initialize list to save acquisition times.
T0 = ['0']

for index in range(len(nombre)):

    # Read file ".rnd".
    archivo = nombre[index] 
    infile = open(archivo, 'r')					

    # Initialize variables for data storage.
    registro, t, T, address, Leq, Lmax, Lmin, LEQT = [], [], [], [], [], [], [], []

    # Read file lines.
    for line in infile:									
        temp = line.split(',')							
        temp = temp[0:]									
        registro.append(temp)							

        # Read words from each line.
        for frase in temp:								

            # Case 1. Set number of records.
            if frase == 'Address':
                address.append(1)

            # Case 2. Saves measurement time.  
            elif frase == 'Start Time':
                t.append(temp[1])

            # Case 3. Save data Leq.   
            elif frase == 'Leq':
                # Convierte string a float.
                num = [float(i) for i in temp[1:-1]]
                Leq.append(num)

            # Case 4. Saves Lmax data.    
            elif frase == 'Lmax':
                # Convierte string a float.
                num = [float(i) for i in temp[1:-1]]
                Lmax.append(num)

            # Case 5. Save data Lmin.   
            elif frase == 'Lmin':
                # Convierte string a float.
                num = [float(i) for i in temp[1:-1]]
                Lmin.append(num);

    # Close file reading.
    infile.close()										

    # Convert lists of data to fixes.
    LeqR = np.array(Leq)
    LmaxR = np.array(Lmax)
    LminR = np.array(Lmin)

    # # Print descriptors by record in file: LeqR, LmaxR, LminR.
    # print 'REGISTROS EN ARCHIVO:', archivo, '\n'
    # print 'LeqR:', LeqR, '\n'
    # print 'LmaxR:', LmaxR, '\n'
    # print 'LminR:', LminR, '\n'    

    # Calculate average descriptors per file.
    LeqF = 10*np.log10((1./LeqR.shape[0])*np.sum(10**(LeqR/10), axis=0))[np.newaxis]
    LmaxF = 10*np.log10((1./LmaxR.shape[0])*np.sum(10**(LmaxR/10), axis=0))[np.newaxis]
    LminF = 10*np.log10((1./LminR.shape[0])*np.sum(10**(LminR/10), axis=0))[np.newaxis]
    T = [t[0], t[-1]]
    d = len(address)

    # Prints average file descriptors: LeqF, LmaxF, LminF.
    print 'DESCRIPTORES PROMEDIO ARCHIVO:', archivo, '\n'
    print 'LeqF:', LeqF, '\n'
    print 'LmaxF:', LmaxF, '\n'
    print 'LminF:', LminF, '\n'
    print 'T:', T, '\n'
    print 'd:', d, '[s]', '\n'

    # Saves average descriptors in total results.
    LEQ[index,0:34] = LeqF
    LMAX[index,0:34] = LmaxF
    LMIN[index,0:34] = LminF
    T0 = T0 + [T]
    D[index] = d

# Print total results: LEQ, LMAX, LMIN.
print 'RESULTADOS:', '\n'
print 'LEQ:', LEQ, '\n'
print 'LMAX:', LMAX, '\n'
print 'LMIN:', LMIN, '\n'
print 'T0:', T0[1:], '\n'
print 'D:', D, '\n'
print 'ESTADO:', 'PROCESADO!', '\n'

# Write results in file.xlsx
# Generates cell names.
alph = list(string.ascii_uppercase)
alph2 = ['A' + index for index in alph]
alph3 = alph + alph2

# Generates labels 1/3 octave bands.
frecuencia = ['Main [dB]', '12.5 Hz', '16 Hz', '20 Hz', '25 Hz', '31.5 Hz', '40 Hz', '50 Hz', '63 Hz', '80 Hz', '100 Hz',
'125 Hz', '160 Hz', '200 Hz', '250 Hz', '315 Hz', '400 Hz', '500 Hz', '630 Hz', '800 Hz', '1 kHz', '1.25 kHz', 
'1.6 kHz', '2 kHz', '2.5 kHz', '3.15 kHz', '4 kHz', '5 kHz', '6.3 kHz', '8 kHz', '10 kHz', '12.5 kHz', '16 kHz', '20 kHz']

# Select data array.
table = LEQ

# Initialize writing workbook.
wb = pyxl.Workbook()
ws = wb.active
ws.title = 'LEQ'

# Generates writing loop in workbook (i: row, j: column).
tableshape = np.shape(table)

print 'Shape', tableshape, '\n'

for i in range(tableshape[0]):
    for j in range(tableshape[1]):    
        # Item - column A.    
        ws['A1'] = 'Item'; 
        ws[alph3[0]+str(i+2)] = i+1

        # File - column B.
        ws['B1'] = 'Archivo'; 
        ws[alph3[1]+str(i+2)] = nombre[i]
        
        # Initial acquisition time - column C.
        ws['C1'] = 'Inicio'; 
        ws[alph3[2]+str(i+2)] = T0[i+1][0]
        
        # Final acquisition time - column D.
        ws['D1'] = 'Fin'; 
        ws[alph3[3]+str(i+2)] = T0[i+1][1]
        
        # Acquisition duration - column E.
        ws['E1'] = 'Duracion [s]'; 
        ws[alph3[4]+str(i+2)] = D[i]
        
        # Frequency - row 1.
        ws[alph3[j+5]+str(1)] = frecuencia[j]
        
        # Data
        ws[alph3[j+5]+str(i+2)] = table[i, j]

# Save information in file.xlsx.
wb.save('results_rion52' + '.xlsx')